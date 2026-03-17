"""datagen/excel_generator.py — GAAP multi-sheet Excel workbook"""
from __future__ import annotations
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

NAVY_FILL = PatternFill('solid', fgColor='1E3A5F')
LIGHT_FILL = PatternFill('solid', fgColor='EBF5FB')
WHITE_FILL = PatternFill('solid', fgColor='FFFFFF')

def _hdr(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    c.fill = NAVY_FILL
    c.alignment = Alignment(horizontal='center', vertical='center')

def _money(ws, row, col, value, bold=False):
    if value is None: return
    c = ws.cell(row=row, column=col, value=round(value, 2))
    c.number_format = '"$"#,##0.00'
    c.font = Font(name='Calibri', size=10, bold=bold)
    c.fill = LIGHT_FILL if row % 2 == 0 else WHITE_FILL

def _pct(ws, row, col, value):
    if value is None: return
    c = ws.cell(row=row, column=col, value=value)
    c.number_format = '0.0%'
    c.font = Font(name='Calibri', size=10)
    c.fill = LIGHT_FILL if row % 2 == 0 else WHITE_FILL

def _lbl(ws, row, text, bold=False):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name='Calibri', size=10, bold=bold, color='1E3A5F' if bold else '000000')

def generate_financial_excel(company, output_path: str) -> None:
    wb = openpyxl.Workbook()
    # ── Sheet 1: Income Statement 3yr ─────────────────────────────────────────
    ws = wb.active; ws.title = "Income Statement"
    ws.column_dimensions['A'].width = 34
    for col, w in [('B',16),('C',16),('D',16)]: ws.column_dimensions[col].width = w
    for ci, h in enumerate(['Line Item','FY 2022','FY 2023','FY 2024'], 1): _hdr(ws,1,ci,h)
    IS_ROWS = [
        ("Revenue","total_revenue",True,False),
        ("Cost of Revenue",lambda f: f["total_revenue"]-f["gross_profit"],False,False),
        ("Gross Profit","gross_profit",True,False),
        ("  Gross Margin %","gross_margin",False,True),
        ("Operating Expenses","operating_expenses",False,False),
        ("Depreciation & Amortization","depreciation_amortization",False,False),
        ("Operating Income (EBIT)","operating_income",True,False),
        ("EBITDA","ebitda",True,False),
        ("  EBITDA Margin %","ebitda_margin",False,True),
        ("Interest Expense","interest_expense",False,False),
        ("Income Before Tax","income_before_tax",True,False),
        ("Income Tax Expense","tax_expense",False,False),
        ("Net Income","net_income",True,False),
        ("  Net Margin %","net_margin",False,True),
    ]
    for ri, (label, key, bold, is_pct) in enumerate(IS_ROWS, 2):
        _lbl(ws, ri, label, bold)
        if key is None: continue
        for ci, fin in enumerate(company.financials, 2):
            val = key(fin) if callable(key) else fin.get(key)
            if is_pct: _pct(ws, ri, ci, val)
            else: _money(ws, ri, ci, val, bold)
    # ── Sheet 2: Balance Sheet ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Balance Sheet")
    ws2.column_dimensions['A'].width = 34
    for col, w in [('B',16),('C',16),('D',16)]: ws2.column_dimensions[col].width = w
    for ci, h in enumerate(['Line Item','Dec 31 2022','Dec 31 2023','Dec 31 2024'], 1): _hdr(ws2,1,ci,h)
    BS_ROWS = [
        ("ASSETS",None,True),("  Cash & Equivalents","cash_and_equivalents",False),
        ("  Accounts Receivable","accounts_receivable",False),("  Inventory","inventory",False),
        ("Total Current Assets","current_assets",True),
        ("  Non-Current Assets",lambda f: f["total_assets"]-f["current_assets"],False),
        ("TOTAL ASSETS","total_assets",True),("","",False),
        ("LIABILITIES",None,True),("  Current Liabilities","current_liabilities",False),
        ("  Long-Term Debt","long_term_debt",False),("TOTAL LIABILITIES","total_liabilities",True),
        ("","",False),("EQUITY",None,True),("TOTAL EQUITY","total_equity",True),
        ("TOTAL LIABILITIES & EQUITY",lambda f: f["total_liabilities"]+f["total_equity"],True),
    ]
    for ri, row in enumerate(BS_ROWS, 2):
        label, key, bold = row
        _lbl(ws2, ri, label, bold)
        if not key: continue
        for ci, fin in enumerate(company.financials, 2):
            val = key(fin) if callable(key) else fin.get(key)
            _money(ws2, ri, ci, val, bold)
    # ── Sheet 3: Key Ratios ────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Key Ratios")
    ws3.column_dimensions['A'].width = 30
    for col, w in [('B',14),('C',14),('D',14)]: ws3.column_dimensions[col].width = w
    for ci, h in enumerate(['Ratio','FY 2022','FY 2023','FY 2024'], 1): _hdr(ws3,1,ci,h)
    RATIOS = [
        ("LIQUIDITY",None,True), ("  Current Ratio","current_ratio",False),
        ("  Cash Ratio",lambda f: f["cash_and_equivalents"]/max(f["current_liabilities"],1),False),
        ("LEVERAGE",None,True), ("  Debt/Equity","debt_to_equity",False),
        ("  Debt/EBITDA","debt_to_ebitda",False), ("  Interest Coverage","interest_coverage_ratio",False),
        ("PROFITABILITY",None,True), ("  Gross Margin","gross_margin",True),
        ("  EBITDA Margin","ebitda_margin",True), ("  Net Margin","net_margin",True),
        ("  Return on Assets",lambda f: f["net_income"]/max(f["total_assets"],1),False),
        ("  Return on Equity",lambda f: f["net_income"]/max(f["total_equity"],1),False),
    ]
    for ri, (label, key, bold) in enumerate(RATIOS, 2):
        _lbl(ws3, ri, label, bold)
        if not key: continue
        for ci, fin in enumerate(company.financials, 2):
            val = key(fin) if callable(key) else fin.get(key)
            if val is None: continue
            if "margin" in label.lower() or "return" in label.lower():
                _pct(ws3, ri, ci, val)
            else:
                c = ws3.cell(row=ri, column=ci, value=round(val, 2))
                c.number_format = '#,##0.00'; c.font = Font(name='Calibri', size=10)
    wb.save(output_path)
